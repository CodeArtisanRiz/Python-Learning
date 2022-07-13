import openpyxl as xl
from openpyxl.chart import BarChart, PieChart, Reference

# def process_workbook(filename):
wb = xl.load_workbook('sheets/transaction_sheet.xlsx')
sheet = wb['Sheet1']
# cell = sheet['a1']
# print(sheet.max_row)
for row in range(2, sheet.max_row + 1):
    # print(row)
    cell = sheet.cell(row, 3)
    print(cell.value)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row, 4)

values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)
bar_chart = BarChart()
bar_chart.add_data(values)
pie_chart = PieChart()
pie_chart.add_data(values)

sheet.add_chart(bar_chart, 'e2')
sheet.add_chart(pie_chart, 'a5')

wb.save('./sheets/processed_transaction_sheet.xlsx')
